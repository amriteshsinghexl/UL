"""
Excel output generator for ULP model results.

Reads a CSV summary file and produces a styled .xlsx workbook with two sheets:

  Sheet 1 "Summary Data"
      One row per projection month (t).  For variables whose formula is a pure
      linear combination of other output columns the cell contains an actual
      Excel formula so that Excel's built-in "Trace Precedents" (blue arrows)
      works natively.  Every cell also carries a comment that shows the
      human-readable actuarial formula.

  Sheet 2 "Formula Reference"
      A lookup table: Variable | Display Name | Formula | Depends On | Part | Description

Usage
-----
    from app.excel_generator import build_excel

    xlsx_bytes = build_excel(csv_path, formula_map)
    with open("output.xlsx", "wb") as f:
        f.write(xlsx_bytes)
"""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

_HEADER_FILL   = "1e40af"   # deep blue  – header row
_FORMULA_FILL  = "dbeafe"   # light blue – cells with Excel formulas
_ALT_ROW_FILL  = "f0f9ff"   # very light – alternating data rows
_PART_COLOURS: dict[str, str] = {
    "Part 2 — Decrements":           "fef3c7",  # amber
    "Part 3 Pass 1 — Cashflows":     "dcfce7",  # green
    "Pass 2 — Backward (Zeroising)": "f3e8ff",  # purple
    "Pass 3 — Forward (Tax & SCR)":  "ffe4e6",  # rose
    "Pass 4 — Backward (Present Values)": "e0f2fe",  # sky
}

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# Column order expected in the CSV / produced by the model
# ---------------------------------------------------------------------------

_COL_ORDER = [
    "t",
    "no_pols_if", "no_pols_ifsm", "no_deaths", "no_surrs", "no_mats",
    "prem_inc_if", "basic_prem_if", "topup_prem_if",
    "op_init_exp_if", "op_ren_exp_if", "invt_exp_if",
    "comm_if", "ovrd_if",
    "death_outgo", "surr_outgo", "mat_outgo", "cog_term_adj",
    "unit_res_bgn", "unit_res_end",
    "unit_inc", "non_unit_inc",
    "cf_before_zv",
    "zeroising_res_if",
    "cf_after_zv", "op_tax", "cf_after_tax",
    "tot_res_if", "solv_cap_req",
    "scr_inv_inc", "scr_inc_tax", "cf_after_scr",
    "pv_cf_after_scr", "pv_prem_inc",
]


def _col_letter(col_map: dict[str, int], name: str) -> str:
    """Return the Excel column letter for *name*, or empty string if absent."""
    idx = col_map.get(name)
    if idx is None:
        return ""
    return get_column_letter(idx)


def _cell_ref(col_map: dict[str, int], name: str, row: int) -> str:
    """Return an absolute Excel cell reference like $B$5."""
    letter = _col_letter(col_map, name)
    if not letter:
        return ""
    return f"${letter}${row}"


# ---------------------------------------------------------------------------
# Excel formula builders
# Variable name → callable(col_map, row, prev_row) → formula string or None
# None means: keep the raw CSV value (no Excel formula possible).
# ---------------------------------------------------------------------------

def _excel_formulas(col_map: dict[str, int]) -> dict[str, Any]:
    """
    Return a dict {variable_name: formula_fn} where formula_fn(row, prev_row)
    produces an Excel formula string (starting with "=").
    prev_row is the row index of the previous time step (may equal row for t=0).
    """

    def _ref(name: str, r: int) -> str:
        return _cell_ref(col_map, name, r)

    formulas: dict[str, Any] = {}

    # no_pols_ifsm[t] = MAX(no_pols_if[t-1] - no_mats[t-1], 0)
    if col_map.get("no_pols_ifsm") and col_map.get("no_pols_if") and col_map.get("no_mats"):
        def _no_pols_ifsm(row: int, prev: int) -> str | None:
            if prev == row:   # t=0 row: keep raw value
                return None
            return f"=MAX({_ref('no_pols_if', prev)}-{_ref('no_mats', prev)},0)"
        formulas["no_pols_ifsm"] = _no_pols_ifsm

    # no_pols_if[t] = no_pols_ifsm - no_deaths - no_surrs - no_mats
    if all(col_map.get(k) for k in ("no_pols_if", "no_pols_ifsm", "no_deaths", "no_surrs", "no_mats")):
        def _no_pols_if(row: int, prev: int) -> str | None:
            if prev == row:
                return None
            return (
                f"={_ref('no_pols_ifsm', row)}"
                f"-{_ref('no_deaths', row)}"
                f"-{_ref('no_surrs', row)}"
                f"-{_ref('no_mats', row)}"
            )
        formulas["no_pols_if"] = _no_pols_if

    # prem_inc_if[t] = basic_prem_if + topup_prem_if
    if all(col_map.get(k) for k in ("prem_inc_if", "basic_prem_if", "topup_prem_if")):
        def _prem_inc(row: int, prev: int) -> str | None:
            return f"={_ref('basic_prem_if', row)}+{_ref('topup_prem_if', row)}"
        formulas["prem_inc_if"] = _prem_inc

    # cf_before_zv — large linear combination of same-row columns
    _cfbzv_deps = [
        "unit_res_bgn", "prem_inc_if", "unit_inc", "non_unit_inc",
        "op_init_exp_if", "op_ren_exp_if", "invt_exp_if",
        "comm_if", "ovrd_if",
        "death_outgo", "surr_outgo", "mat_outgo", "cog_term_adj",
        "unit_res_end",
    ]
    if all(col_map.get(k) for k in _cfbzv_deps) and col_map.get("cf_before_zv"):
        def _cf_before_zv(row: int, prev: int) -> str | None:
            add = ["unit_res_bgn", "prem_inc_if", "unit_inc", "non_unit_inc"]
            sub = [
                "op_init_exp_if", "op_ren_exp_if", "invt_exp_if",
                "comm_if", "ovrd_if",
                "death_outgo", "surr_outgo", "mat_outgo", "cog_term_adj",
                "unit_res_end",
            ]
            pos = "+".join(_ref(k, row) for k in add)
            neg = "-".join(_ref(k, row) for k in sub)
            return f"={pos}-{neg}"
        formulas["cf_before_zv"] = _cf_before_zv

    # cf_after_tax = cf_after_zv - op_tax
    if all(col_map.get(k) for k in ("cf_after_tax", "cf_after_zv", "op_tax")):
        def _cf_after_tax(row: int, prev: int) -> str | None:
            return f"={_ref('cf_after_zv', row)}-{_ref('op_tax', row)}"
        formulas["cf_after_tax"] = _cf_after_tax

    # tot_res_if = unit_res_end + zeroising_res_if
    if all(col_map.get(k) for k in ("tot_res_if", "unit_res_end", "zeroising_res_if")):
        def _tot_res(row: int, prev: int) -> str | None:
            return f"={_ref('unit_res_end', row)}+{_ref('zeroising_res_if', row)}"
        formulas["tot_res_if"] = _tot_res

    # cf_after_scr = cf_after_tax + scr_prev - scr + scr_inv_inc - scr_inc_tax
    if all(col_map.get(k) for k in ("cf_after_scr", "cf_after_tax", "solv_cap_req", "scr_inv_inc", "scr_inc_tax")):
        def _cf_after_scr(row: int, prev: int) -> str | None:
            if prev == row:
                return None
            return (
                f"={_ref('cf_after_tax', row)}"
                f"+{_ref('solv_cap_req', prev)}"
                f"-{_ref('solv_cap_req', row)}"
                f"+{_ref('scr_inv_inc', row)}"
                f"-{_ref('scr_inc_tax', row)}"
            )
        formulas["cf_after_scr"] = _cf_after_scr

    return formulas


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _hfill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF", size=9)


def _data_font(bold: bool = False) -> Font:
    return Font(bold=bold, size=9)


def _add_comment(cell: Any, text: str, author: str = "ULP Model") -> None:
    c = Comment(text, author)
    c.width  = 320
    c.height = max(80, 20 * text.count("\n") + 40)
    cell.comment = c


# ---------------------------------------------------------------------------
# Sheet 1: Summary Data
# ---------------------------------------------------------------------------

def _build_summary_sheet(
    ws: Any,
    rows: list[dict[str, str]],
    col_order: list[str],
    formula_map: dict[str, dict],
) -> None:
    if not _HAS_OPENPYXL:
        return

    # ---- Build column index map: variable_name → excel column index (1-based) ----
    col_map: dict[str, int] = {}
    for ci, name in enumerate(col_order, start=1):
        col_map[name] = ci

    excel_fns = _excel_formulas(col_map)

    # ---- Header row ----
    for ci, name in enumerate(col_order, start=1):
        cell = ws.cell(row=1, column=ci)
        entry = formula_map.get(name)
        cell.value = entry["display_name"] if entry else name
        cell.font  = _header_font()
        cell.fill  = _hfill(_HEADER_FILL)
        cell.border = _BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(cell.value or "") + 2)

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "B2"

    # ---- Data rows ----
    for ri, row_data in enumerate(rows, start=2):
        prev_row = ri - 1 if ri > 2 else ri   # row 2 is t=0; no "previous" for t=0

        for ci, name in enumerate(col_order, start=1):
            cell = ws.cell(row=ri, column=ci)

            raw_val = row_data.get(name, "")
            try:
                numeric = float(raw_val)
            except (ValueError, TypeError):
                numeric = None

            # Decide: Excel formula or raw value?
            formula_fn = excel_fns.get(name)
            if formula_fn is not None:
                formula = formula_fn(ri, prev_row)
            else:
                formula = None

            if formula:
                cell.value = formula
                cell.fill  = _hfill(_FORMULA_FILL)
            else:
                cell.value = numeric if numeric is not None else raw_val
                fill_hex = _ALT_ROW_FILL if ri % 2 == 0 else "FFFFFF"
                cell.fill = _hfill(fill_hex)

            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="right", vertical="center")

            if numeric is not None and name != "t":
                cell.number_format = "#,##0.00"

            # Cell comment: actuarial formula
            entry = formula_map.get(name)
            if entry:
                comment_text = (
                    f"Variable: {entry['name']}\n"
                    f"Stage: {entry['part']}\n\n"
                    f"Formula:\n  {entry['formula']}\n\n"
                    f"Depends on:\n  {', '.join(entry['depends_on']) if entry['depends_on'] else '(inputs/parameters)'}\n\n"
                    f"Description:\n  {entry['description']}"
                )
                _add_comment(cell, comment_text)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_order))}1"


# ---------------------------------------------------------------------------
# Sheet 2: Formula Reference
# ---------------------------------------------------------------------------

def _build_formula_sheet(
    ws: Any,
    formula_map: dict[str, dict],
    col_order: list[str],
) -> None:
    if not _HAS_OPENPYXL:
        return

    headers = ["Variable", "Display Name", "Stage", "Formula", "Depends On", "Description"]
    col_widths = [18, 28, 30, 60, 35, 60]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value = h
        cell.font  = _header_font()
        cell.fill  = _hfill(_HEADER_FILL)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    # Group by part
    part_order: list[str] = []
    by_part: dict[str, list[dict]] = {}
    for name in col_order:
        entry = formula_map.get(name)
        if not entry:
            continue
        part = entry["part"]
        if part not in by_part:
            by_part[part] = []
            part_order.append(part)
        by_part[part].append(entry)

    ri = 2
    for part in part_order:
        part_fill = _hfill(_PART_COLOURS.get(part, "f1f5f9"))
        for entry in by_part[part]:
            cells = [
                entry["name"],
                entry["display_name"],
                part,
                entry["formula"],
                ", ".join(entry["depends_on"]),
                entry["description"],
            ]
            for ci, val in enumerate(cells, start=1):
                cell = ws.cell(row=ri, column=ci)
                cell.value = val
                cell.font  = _data_font()
                cell.fill  = part_fill
                cell.border = _BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[ri].height = 42
            ri += 1


# ---------------------------------------------------------------------------
# Sheet 3: Python Source (AST-extracted)
# ---------------------------------------------------------------------------

def _build_source_sheet(
    ws: Any,
    formula_map: dict[str, dict],
    col_order: list[str],
) -> None:
    if not _HAS_OPENPYXL:
        return

    headers = ["Variable", "Python Source (from model)"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value = h
        cell.font  = _header_font()
        cell.fill  = _hfill(_HEADER_FILL)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100
    ws.row_dimensions[1].height = 22

    ri = 2
    for name in col_order:
        entry = formula_map.get(name)
        if not entry:
            continue
        src = entry.get("python_source") or "(not extracted)"
        ws.cell(row=ri, column=1).value = name
        ws.cell(row=ri, column=1).font  = Font(bold=True, size=9)
        ws.cell(row=ri, column=1).border = _BORDER

        src_cell = ws.cell(row=ri, column=2)
        src_cell.value = src
        src_cell.font  = Font(name="Courier New", size=8)
        src_cell.border = _BORDER
        src_cell.alignment = Alignment(wrap_text=True, vertical="top")
        line_count = src.count("\n") + 1
        ws.row_dimensions[ri].height = max(18, min(line_count * 14, 120))
        ri += 1


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_excel(
    csv_path: Path | str,
    formula_map: dict[str, dict],
    selected_fields: list[str] | None = None,
) -> bytes:
    """
    Build an .xlsx workbook from *csv_path* and return the raw bytes.

    Parameters
    ----------
    csv_path        : path to a summary_scen*.csv file
    formula_map     : {variable_name: formula_entry_dict} from formula_extractor
    selected_fields : if provided, only include these columns (plus "t") in the
                      Summary Data sheet; all columns are still shown in the
                      Formula Reference and Python Source sheets.
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is required for Excel generation. "
            "Install it with: pip install openpyxl"
        )

    csv_path = Path(csv_path)
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        csv_headers = reader.fieldnames or []
        rows = list(reader)

    # Build the column order: prefer _COL_ORDER, then any extras from the CSV
    col_order: list[str] = []
    seen: set[str] = set()
    for name in _COL_ORDER:
        if name in csv_headers:
            col_order.append(name)
            seen.add(name)
    for name in csv_headers:
        if name not in seen:
            col_order.append(name)

    # Apply field selection — always keep "t" (the time index column)
    if selected_fields:
        allowed = set(selected_fields) | {"t"}
        col_order = [c for c in col_order if c in allowed]

    wb = openpyxl.Workbook()

    # Sheet 1: Summary Data
    ws_data = wb.active
    ws_data.title = "Summary Data"
    _build_summary_sheet(ws_data, rows, col_order, formula_map)

    # Sheet 2: Formula Reference
    ws_ref = wb.create_sheet("Formula Reference")
    _build_formula_sheet(ws_ref, formula_map, col_order)

    # Sheet 3: Python Source
    ws_src = wb.create_sheet("Python Source (AST)")
    _build_source_sheet(ws_src, formula_map, col_order)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
